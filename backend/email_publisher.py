#!/usr/bin/env python3
"""
email_publisher.py
==================
Lee correos en la casilla configurada, detecta adjuntos Excel (.xlsx / .xls)
y publica automáticamente cada fila como un producto en OkVenta/MercadoFoto.

Uso:
    python email_publisher.py                  # bucle continuo
    python email_publisher.py --once           # procesa una sola vez y sale
    python email_publisher.py --config ruta.json

Configuración: config_email.json (mismo directorio)
"""

import argparse
import imaplib
import email
import io
import json
import os
import sys
import time
import traceback
from email.header import decode_header
from pathlib import Path
from typing import Optional

import openpyxl
import requests

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
DEFAULT_CONFIG = BASE_DIR / "config_email.json"


def cargar_config(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)
    required = ["imap_server", "imap_port", "email", "password", "api_url"]
    for key in required:
        if key not in cfg:
            raise ValueError(f"Falta '{key}' en {path}")
    cfg.setdefault("check_interval_seconds", 300)
    cfg.setdefault("marca_leido", True)
    cfg.setdefault("carpeta", "INBOX")
    return cfg


# ─────────────────────────────────────────────
# DETECCIÓN FLEXIBLE DE COLUMNAS
# ─────────────────────────────────────────────

# Para cada campo destino, lista de posibles nombres de cabecera (en minúsculas)
COLUMN_VARIANTS: dict[str, list[str]] = {
    "titulo":      ["titulo", "título", "nombre", "producto", "name", "title"],
    "precio":      ["precio", "price", "valor", "costo", "monto"],
    "descripcion": ["descripcion", "descripción", "description", "detalle", "desc", "observaciones"],
    "categoria":   ["categoria", "categoría", "category", "tipo", "rubro"],
    "subcategoria":["subcategoria", "subcategoría", "subcategory", "subtipo", "subrubro"],
    "user_id":     ["user_id", "userid", "usuario_id", "id_usuario", "id", "usuario"],
    "imagen":      ["imagen", "image", "foto", "photo", "archivo", "file", "imagen1", "foto1"],
    "imagen2":     ["imagen2", "foto2", "image2", "photo2"],
    "imagen3":     ["imagen3", "foto3", "image3", "photo3"],
    "lat":         ["lat", "latitud", "latitude"],
    "lng":         ["lng", "lon", "longitud", "longitude"],
}


def _normalizar(texto: str) -> str:
    """Minúsculas, sin espacios extra, sin tildes problemáticas."""
    return str(texto).strip().lower().replace("  ", " ")


def detectar_mapa_columnas(cabeceras: list[str]) -> dict[str, int]:
    """
    Dado el listado de cabeceras de la hoja, devuelve un dict
    {campo_destino: índice_columna}.
    """
    mapa: dict[str, int] = {}
    for idx, cab in enumerate(cabeceras):
        cab_norm = _normalizar(cab)
        for campo, variantes in COLUMN_VARIANTS.items():
            if campo not in mapa and cab_norm in variantes:
                mapa[campo] = idx
                break
    return mapa


def buscar_fila_cabecera(ws) -> Optional[int]:
    """
    Recorre las primeras 10 filas buscando aquella que contenga al menos
    'titulo'/'título'/'nombre' Y 'precio'. Devuelve el número de fila (1-based).
    """
    imprescindibles = set(COLUMN_VARIANTS["titulo"]) | set(COLUMN_VARIANTS["precio"])
    for row_num in range(1, 11):
        valores = [_normalizar(str(ws.cell(row=row_num, column=c).value or ""))
                   for c in range(1, ws.max_column + 1)]
        if any(v in imprescindibles for v in valores):
            return row_num
    return None


# ─────────────────────────────────────────────
# LECTURA DEL EXCEL
# ─────────────────────────────────────────────

def leer_productos_excel(
    contenido_excel: bytes,
    imagenes_adjuntas: dict[str, bytes],
) -> list[dict]:
    """
    Parsea el Excel y devuelve una lista de dicts con los campos del producto.
    `imagenes_adjuntas` es {nombre_archivo_lower: bytes} de las imágenes del correo.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contenido_excel), data_only=True)
    ws = wb.active

    fila_cab = buscar_fila_cabecera(ws)
    if fila_cab is None:
        print("  ⚠  No se detectó fila de cabeceras en el Excel.")
        return []

    # Leer cabeceras
    cabeceras = [
        str(ws.cell(row=fila_cab, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ]
    mapa = detectar_mapa_columnas(cabeceras)
    print(f"  Mapa de columnas detectado: {mapa}")

    if "titulo" not in mapa or "precio" not in mapa:
        print("  ⚠  No se encontraron columnas obligatorias (titulo, precio).")
        return []

    productos = []
    for row_num in range(fila_cab + 1, ws.max_row + 1):
        fila = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]

        # Ignorar filas vacías
        if all(v is None or str(v).strip() == "" for v in fila):
            continue

        def get(campo: str, default=None):
            if campo in mapa:
                val = fila[mapa[campo]]
                return val if val is not None else default
            return default

        titulo = str(get("titulo", "")).strip()
        if not titulo:
            continue

        try:
            precio = float(str(get("precio", 0)).replace(",", ".").replace("$", "").strip())
        except (ValueError, TypeError):
            print(f"  ⚠  Fila {row_num}: precio inválido, se omite.")
            continue

        descripcion = str(get("descripcion", titulo)).strip() or titulo
        categoria   = str(get("categoria",  "")).strip() or None
        subcategoria= str(get("subcategoria","")).strip() or None
        user_id_raw = get("user_id")
        lat_raw     = get("lat")
        lng_raw     = get("lng")

        user_id = None
        if user_id_raw is not None:
            try:
                user_id = int(float(str(user_id_raw)))
            except (ValueError, TypeError):
                pass

        lat = None
        lng = None
        try:
            if lat_raw is not None:
                lat = float(str(lat_raw).replace(",", "."))
            if lng_raw is not None:
                lng = float(str(lng_raw).replace(",", "."))
        except (ValueError, TypeError):
            pass

        # Imágenes: intentar emparejar por nombre de archivo
        imagenes_bytes: list[bytes] = []
        for campo_img in ["imagen", "imagen2", "imagen3"]:
            nombre_img = str(get(campo_img, "")).strip().lower()
            if nombre_img and nombre_img in imagenes_adjuntas:
                imagenes_bytes.append(imagenes_adjuntas[nombre_img])

        # Si no hay imágenes referenciadas, tomar las adjuntas en orden
        if not imagenes_bytes and imagenes_adjuntas:
            imagenes_bytes = list(imagenes_adjuntas.values())[:4]

        productos.append({
            "titulo":       titulo,
            "descripcion":  descripcion,
            "precio":       precio,
            "categoria":    categoria,
            "subcategoria": subcategoria,
            "user_id":      user_id,
            "lat":          lat,
            "lng":          lng,
            "imagenes":     imagenes_bytes,  # lista de bytes
        })

    return productos


# ─────────────────────────────────────────────
# PUBLICAR VÍA API
# ─────────────────────────────────────────────

def publicar_producto(api_url: str, producto: dict) -> bool:
    """
    Llama a POST /publicar con multipart/form-data.
    Devuelve True si tuvo éxito.
    """
    if not producto["imagenes"]:
        print(f"  ⚠  '{producto['titulo']}': sin imagen, se omite.")
        return False

    if not producto["user_id"]:
        print(f"  ⚠  '{producto['titulo']}': sin user_id, se omite.")
        return False

    campos = {
        "titulo":      (None, str(producto["titulo"])),
        "descripcion": (None, str(producto["descripcion"])),
        "precio":      (None, str(producto["precio"])),
        "user_id":     (None, str(producto["user_id"])),
    }
    if producto["categoria"]:
        campos["categoria"]    = (None, producto["categoria"])
    if producto["subcategoria"]:
        campos["subcategoria"] = (None, producto["subcategoria"])
    if producto["lat"] is not None:
        campos["lat"] = (None, str(producto["lat"]))
    if producto["lng"] is not None:
        campos["lng"] = (None, str(producto["lng"]))

    # Construir archivos: file (obligatorio) + file2, file3, file4
    slots = ["file", "file2", "file3", "file4"]
    archivos = {}
    for i, img_bytes in enumerate(producto["imagenes"][:4]):
        archivos[slots[i]] = (f"imagen{i+1}.jpg", img_bytes, "image/jpeg")

    try:
        resp = requests.post(
            f"{api_url}/publicar",
            data={k: v[1] for k, v in campos.items()},
            files=archivos,
            timeout=120,
        )
        if resp.status_code == 200:
            data = resp.json()
            print(f"  ✓  Publicado: '{producto['titulo']}' → {data.get('imagen_url','')}")
            return True
        else:
            print(f"  ✗  Error {resp.status_code}: {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"  ✗  Excepción al publicar '{producto['titulo']}': {e}")
        return False


# ─────────────────────────────────────────────
# LECTURA DE CORREOS
# ─────────────────────────────────────────────

def obtener_nombre_adjunto(part) -> str:
    """Decodifica el nombre del adjunto de forma segura."""
    filename = part.get_filename() or ""
    if filename:
        decoded_parts = decode_header(filename)
        partes = []
        for data, charset in decoded_parts:
            if isinstance(data, bytes):
                partes.append(data.decode(charset or "utf-8", errors="replace"))
            else:
                partes.append(data)
        filename = "".join(partes)
    return filename


def procesar_correo(msg, api_url: str) -> int:
    """
    Extrae adjuntos de un correo y publica los productos encontrados.
    Devuelve el número de productos publicados.
    """
    # Separar adjuntos Excel e imágenes
    excels: list[tuple[str, bytes]] = []          # (nombre, bytes)
    imagenes: dict[str, bytes] = {}               # {nombre_lower: bytes}

    for part in msg.walk():
        disposition = part.get_content_disposition() or ""
        content_type = part.get_content_type()
        filename = obtener_nombre_adjunto(part)
        filename_lower = filename.lower()

        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        # Excel
        if (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls") or
                "spreadsheet" in content_type or "excel" in content_type):
            excels.append((filename, payload))

        # Imagen
        elif (content_type.startswith("image/") or
              any(filename_lower.endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"])):
            imagenes[filename_lower] = payload

    if not excels:
        return 0

    total_publicados = 0
    for nombre_excel, contenido_excel in excels:
        print(f"  Procesando Excel: {nombre_excel}")
        try:
            productos = leer_productos_excel(contenido_excel, imagenes)
            print(f"  → {len(productos)} producto(s) encontrado(s)")
            for p in productos:
                if publicar_producto(api_url, p):
                    total_publicados += 1
        except Exception as e:
            print(f"  ✗  Error procesando {nombre_excel}: {e}")
            traceback.print_exc()

    return total_publicados


def revisar_correos(cfg: dict) -> int:
    """
    Conecta al IMAP, procesa correos no leídos con adjuntos Excel.
    Devuelve el total de productos publicados.
    """
    server   = cfg["imap_server"]
    port     = cfg["imap_port"]
    usuario  = cfg["email"]
    password = cfg["password"]
    carpeta  = cfg["carpeta"]
    api_url  = cfg["api_url"].rstrip("/")
    marcar   = cfg["marca_leido"]

    print(f"\n[{time.strftime('%Y-%m-%d %H:%M:%S')}] Revisando {usuario}…")

    try:
        mail = imaplib.IMAP4_SSL(server, port)
        mail.login(usuario, password)
        mail.select(carpeta)

        # Buscar no leídos con adjunto (IMAP no filtra por adjunto nativamente,
        # pero filtramos tras leer el mensaje)
        status, data = mail.search(None, "UNSEEN")
        if status != "OK":
            print("  No se pudo buscar mensajes.")
            mail.logout()
            return 0

        ids = data[0].split()
        if not ids:
            print("  Sin mensajes nuevos.")
            mail.logout()
            return 0

        print(f"  {len(ids)} mensaje(s) sin leer.")
        total = 0

        for uid in ids:
            res, msg_data = mail.fetch(uid, "(RFC822)")
            if res != "OK":
                continue

            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            asunto_raw = msg.get("Subject", "")
            decoded = decode_header(asunto_raw)
            asunto = ""
            for data, charset in decoded:
                if isinstance(data, bytes):
                    asunto += data.decode(charset or "utf-8", errors="replace")
                else:
                    asunto += data
            print(f"\n  ✉  Asunto: {asunto}")

            publicados = procesar_correo(msg, api_url)
            total += publicados

            if marcar and publicados > 0:
                mail.store(uid, "+FLAGS", "\\Seen")
                print(f"  → Marcado como leído ({publicados} publicado(s)).")
            elif publicados == 0:
                print("  → Sin productos publicados (sin Excel válido o sin datos).")

        mail.logout()
        return total

    except imaplib.IMAP4.error as e:
        print(f"  ✗  Error IMAP: {e}")
        return 0
    except Exception as e:
        print(f"  ✗  Error inesperado: {e}")
        traceback.print_exc()
        return 0


# ─────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Email auto-publisher para OkVenta")
    parser.add_argument("--once",   action="store_true", help="Procesar una vez y salir")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Ruta al JSON de configuración")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        print(f"✗ No se encontró la configuración: {config_path}")
        print("  Crea config_email.json con: imap_server, imap_port, email, password, api_url")
        sys.exit(1)

    cfg = cargar_config(config_path)
    intervalo = cfg["check_interval_seconds"]

    print("=" * 55)
    print("  OkVenta — Email Auto-Publisher")
    print(f"  Cuenta  : {cfg['email']}")
    print(f"  Servidor: {cfg['imap_server']}:{cfg['imap_port']}")
    print(f"  API     : {cfg['api_url']}")
    print(f"  Intervalo: {intervalo}s" if not args.once else "  Modo: una sola ejecución")
    print("=" * 55)

    if args.once:
        revisar_correos(cfg)
        return

    # Bucle continuo
    while True:
        try:
            revisar_correos(cfg)
        except KeyboardInterrupt:
            print("\nDetenido por el usuario.")
            break
        except Exception as e:
            print(f"Error en bucle principal: {e}")
            traceback.print_exc()

        print(f"  Próxima revisión en {intervalo}s… (Ctrl+C para salir)\n")
        try:
            time.sleep(intervalo)
        except KeyboardInterrupt:
            print("\nDetenido por el usuario.")
            break


if __name__ == "__main__":
    main()
