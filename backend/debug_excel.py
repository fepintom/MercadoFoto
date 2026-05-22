#!/usr/bin/env python3
"""
debug_excel.py — lee el último correo con Excel y muestra qué ve openpyxl.
"""
import imaplib, email, io, json
from pathlib import Path
import openpyxl

cfg = json.loads((Path(__file__).parent / "config_email.json").read_text())
mail = imaplib.IMAP4_SSL(cfg["imap_server"], int(cfg.get("imap_port", 993)))
mail.login(cfg["email"], cfg["password"])
mail.select("INBOX")

_, data = mail.search(None, "ALL")
uids = data[0].split()
print(f"Total correos: {len(uids)}")

# Buscar el último correo con Excel adjunto
for uid in reversed(uids[-10:]):
    _, msg_data = mail.fetch(uid, "(RFC822)")
    msg = email.message_from_bytes(msg_data[0][1])
    print(f"\nUID {uid.decode()} | De: {msg.get('From')} | Asunto: {msg.get('Subject')}")

    for part in msg.walk():
        fn = part.get_filename() or ""
        if fn.lower().endswith((".xlsx", ".xls")):
            print(f"  Excel encontrado: {fn}")
            contenido = part.get_payload(decode=True)
            wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
            print(f"  Hojas: {wb.sheetnames}")
            ws = wb.active
            print(f"  Hoja activa: {ws.title}")
            print(f"  max_row={ws.max_row}, max_column={ws.max_column}")
            print(f"  Primeras 8 filas:")
            for row_num in range(1, min(ws.max_row + 1, 9)):
                fila = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
                # Mostrar solo filas con algo
                if any(v is not None for v in fila):
                    print(f"    Fila {row_num}: {fila}")
            break
    else:
        continue
    break

mail.logout()
